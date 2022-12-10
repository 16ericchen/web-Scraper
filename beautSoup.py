from bs4 import BeautifulSoup
from selenium import webdriver
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from openpyxl import load_workbook,Workbook
import os
findChip = {}
oemTrade = {}
octoParts = {}
mfrDict = {}
apnDict = {}
listOfDicts = [oemTrade,octoParts,findChip]
validDistributors = ["Digi-Key",'Arrow Electronics','Future Electronics','Verical','Newark','Mouser Electronics','Mouser','Master','Rochester','Master Electronics']
findChipDistributors = ["Digi-Key",'Arrow Electronics','Future Electronics','Verical','Newark','Mouser Electronics']
os.chdir(os.path.dirname(__file__))
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    startRow = 1
    mfrPartNumber = []

    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    for row in sheet.iter_rows(min_col=1,max_col=1,min_row=1,max_row=20):
        if row[0].value == "Level":
            startRow = row[0].row
            break
    for col in sheet.iter_cols(min_col=10,min_row=startRow,max_row=sheet.max_row):
        if col[0].value == 'Mfr Name':
            for i in range(1,len(col)):
                if col[i].value != None:
                    mfrDict[i] = col[i].value
        if col[0].value == 'Mfr. Part Number':
            for i in range(1,len(col)):
                if sheet.cell(row=startRow+i,column=2).value not in apnDict.keys():
                    apnDict[sheet.cell(row=startRow+i,column=2).value] = []
                if col[i].value != None:
                    if i in mfrDict.keys():     
                        if 'YAGEO' in mfrDict[i]:
                            print('Add Part Number To Database')
                        elif 'Cyntec' in mfrDict[i]:
                            print('Add Part To Cyntec List')
                        else:
                            mfrPartNumber.append(col[i].value)
                            apnDict[sheet.cell(row=startRow+i,column=2).value].append(col[i].value)
            mfrDict.clear()
    for x in apnDict:
        for y in apnDict[x]:
            searchSites(y)
    createXL()

def createXL():
    print('here')
    wx = Workbook()
    log = wx.active
    rnum = 1
    for x in apnDict:
        log.cell(row = rnum,column=1,value = x)
        for y in apnDict[x]:
            print(y)
            log.cell(row=rnum+1,column=1,value= y)
            for z in oemTrade[y]:
                for h in range(len(z)):
                    print(z[h])
                    print(type(z[h]))
                    if type(z[h]) is list:
                        log.cell(row=rnum+2,column=1,value = z[h][0][0])
                        log.cell(row=rnum+3,column=1,value = z[h][0][1])
                        rnum+=2
                    else: 
                        log.cell(row=rnum+2,column=1,value = z[h])
                        rnum+=1
                rnum+=1
            rnum+=1
    wx.save(filename='bom.xlsx')
def searchSites(partNumber):    
    
    options = Options()
    options.add_argument('--headless')
    urls = ["https://www.oemstrade.com/search/"+str(partNumber),"https://octopart.com/search?q="+str(partNumber)+"&currency=USD&specs=0","https://www.findchips.com/search/"+str(partNumber)]
    functions = [oemTrades,octoPart,findChips]
    
    for x in range(len(urls)):
        driver = webdriver.Firefox(options = options)
        driver.get(urls[x])
        print(urls[x])
        # returns list [Manufacturer,Part Number,sock,prices]
        listOfDicts[x][partNumber] = functions[x](BeautifulSoup(driver.page_source, "html.parser"),driver)
        driver.quit()

def oemTrades(soup,driver):
    x = soup.findAll('div', class_='distributor-results')
    results = []
    for y in x:
        if y.get('data-distributorname') in findChipDistributors:
            z = []
            tdStock = y.find('td',class_='td-stock')
            tdPN = y.find('td',class_='td-part-number')
            tdPrice = y.find('td',class_='td-price')
            tdManufacturer = y.find('td',class_='td-distributor-name')
            stock = re.sub("[^0-9]", "", tdStock.get_text().replace("\n", ""))
            test = tdPrice.findAll('li')
            for x in test:
                z.append(x.get_text().split(' '))
            results.append([tdManufacturer.get_text().replace("\n", "").strip(),y.get('data-distributor_name'),tdPN.find('a').get_text(),stock,z])
    return results

def findChips(soup,driver):
    x = soup.findAll('div', class_='distributor-results')
    results = []
    for y in x:
        if y.get('data-distributor_name') in validDistributors:
            z=[]
            tdStock = y.find('td',class_='td-stock')
            tdPrice = y.find('td',class_='td-price')
            tdMNF = y.find('td',class_='td-mfg')
            tdPN = y.find('div',class_="part-name").find('a',class_='onclick').get_text().strip()
            stock = re.sub("[^0-9]", "", tdStock.get_text().replace("\n", ""))
            test = tdPrice.findAll('li')
            for x in test:
                z.append(x.get_text().replace("\n",'').strip().replace(' ','').split('$'))
            results.append([tdMNF.get_text().strip(),y.get('data-distributor_name'),tdPN,stock,z])
    return results

def octoPart(soup,driver):
    try:
        driver.find_element(By.XPATH, '//button[text()="Show All"]').click()
    except:
        pass
    finally:
        x = soup.findAll('tbody')
        results = [] 
        if x:
            for z in x[0]:
                info = []
                t = z.findAll('td')
                result = ''.join([i for i in t[1].get_text() if not i.isdigit()])
                if result in validDistributors:
                    for n in range(len(t)):
                        info.append(t[n].get_text())
                results.append(info)
        return results
flag = False
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()












# options = Options()
# options.add_argument('--headless')
# findChipDistributors = ["Digi-Key",'Arrow Electronics','Future Electronics','Verical','Newark','Mouser Electronics','Mouser','Master','Rochester']
# driver = webdriver.Firefox(options = options)
# stack = []
# validDistributors = ["Digi-Key",'Arrow Electronics','Future Electronics','Verical','Newark','Mouser Electronics','Mouser','Master','Rochester','Master Electronics']
# print('here')
# url = "https://octopart.com/search?q=GRM188R6YA106MA73&currency=USD&specs=0"
# driver.get(url)

# t=driver.find_element(By.XPATH, '//button[text()="Show All"]').click()
# soup = BeautifulSoup(driver.page_source, "html.parser")

# try:
#     driver.find_element(By.XPATH, '//button[text()="Show All"]').click()
# except:
#     pass
# finally:
#     print('here in OctoPart')
#     x = soup.findAll('tbody')
#     results = [] 
#     if x:
        
#         for z in x[0]:
#             info = []
#             t = z.findAll('td')
#             result = ''.join([i for i in t[1].get_text() if not i.isdigit()])
#             if result in validDistributors:
#                 for n in range(len(t)):
#                     info.append(t[n].get_text())
#             print('info',info)
#             results.append(info)
#     print(results)


# url = "https://www.findchips.com/search/5PB1106CMGI"
# driver.get(url)


# soup = BeautifulSoup(driver.page_source, "html.parser")
# x = soup.findAll('div', class_='distributor-results')
# for y in x:
#     if y.get('data-distributor_name') in distributors:
#         z=[]
#         print(y.get('data-distributor_name'))
#         tdStock = y.find('td',class_='td-stock')
#         tdPrice = y.find('td',class_='td-price')
#         tdMNF = y.find('td',class_='td-mfg')
#         tdPN = y.find('div',class_="part-name").find('a',class_='onclick').get_text().strip()
#         stock = re.sub("[^0-9]", "", tdStock.get_text().replace("\n", ""))
#         test = tdPrice.findAll('li')
#         # print(tdPN)
#         for x in test:
#             z.append(x.get_text().replace("\n",'').strip().replace(' ','').split('$'))
#         print(tdPN,stock,z,y.get('data-distributor_name'),tdMNF.get_text().strip())

# url = "https://www.oemstrade.com/search/GRM0225C1C221JA02"
# driver.get(url)

# soup = BeautifulSoup(driver.page_source, "html.parser")
# x = soup.findAll('div', class_='distributor-results')
# for y in x:
#     if y.get('data-distributorname') in findChipDistributors:
#         z = []
#         print(y.get('data-distributorname'))
#         tdStock = y.find('td',class_='td-stock')
#         tdPN = y.find('td',class_='td-part-number')
#         tdPrice = y.find('td',class_='td-price')
#         tdManufacturer = y.find('td',class_='td-distributor-name')
#         stock = re.sub("[^0-9]", "", tdStock.get_text().replace("\n", ""))
#         test = tdPrice.findAll('li')
#         for x in test:
#             z.append(x.get_text().split(' '))
#         # distributors = tdDistributor.get_text().replace("\n", "")
#         print('stock:',stock,'Manufacturer:',tdManufacturer.get_text().replace("\n", ""),' PN: ',tdPN.find('a').get_text(),'distributor: ', y.get('data-distributorname'))



# oemstrade format: https://www.oemstrade.com/search/GRM0225C1C221JA02
# octopart format: https://octopart.com/search?q=GRM0225C1C221JA02&currency=USD&specs=0
# finchips format: https://www.findchips.com/search/GRM0225C1C221JA02

